from itertools import product

from openpyxl.styles import Alignment

from src.app.DB import *
from openpyxl import load_workbook
import openpyxl


def enumeration(name_file, water_expenditure, concentration, money_max):
    wb = openpyxl.Workbook()
    wb.save(name_file)
    stroka_pls = 1
    need_concent = concentration
    device_name, device_list, insoluble_list = get_all()
    spisok_variations = list()
    spisok_posled = list(product(device_name[0], device_name[1], device_name[2], device_name[3],
                       device_name[4], device_name[5], device_name[6], device_name[7]))
    spisok_posled1 = list(product(device_name[1], device_name[2], device_name[3],
                       device_name[4], device_name[5], device_name[6], device_name[7]))
    spisok_posled2 = list(product(device_name[2], device_name[3],device_name[4], device_name[5], device_name[6],
                                  device_name[7]))
    spisok_posled3 = list(product(device_name[3], device_name[4], device_name[5], device_name[6], device_name[7]))
    spisok_posled4 = list(product(device_name[4], device_name[5], device_name[6], device_name[7]))
    spisok_posled5 = list(product(device_name[5], device_name[6], device_name[7]))
    spisok_posled6 = list(product(device_name[6], device_name[7]))
    spisok_posled7 = list(product(device_name[7]))
    spisok_posled.extend(spisok_posled1)
    spisok_posled.extend(spisok_posled2)
    spisok_posled.extend(spisok_posled3)
    spisok_posled.extend(spisok_posled4)
    spisok_posled.extend(spisok_posled5)
    spisok_posled.extend(spisok_posled6)
    spisok_posled.extend(spisok_posled7)

    money = 0
    money_nice = money_max
    spisok_nice = list()


    for i in range(len(spisok_posled)):
        money = 0
        concent_changes = insoluble_list.get('Взвешанные')
        for j in range(len(spisok_posled[i])):
            for l in range(len(device_list)):
                if device_list[l].get('name_device') == spisok_posled[i][j] and device_list[l].get('start') >= concent_changes and device_list[l].get('expenditure') >= water_expenditure:
                    if concent_changes < float(need_concent):
                        break
                    else:
                        money = money + device_list[l].get('cost')
                        start = concent_changes
                        concent_changes = round(effectiveness(concent_changes, device_list[l].get('effectiveness')), 2)
                        spisok_variations.append({'name': device_list[l].get('name_device'), 'start': start,
                                                  'effectiv': device_list[l].get('effectiveness'), 'end': concent_changes,
                                                  'end/PDK': concent_changes/float(need_concent), 'cost': device_list[l].get('cost')})

        if len(spisok_variations) != 0:
            if money <= money_max:
                if money < money_nice:
                    spisok_nice.clear()
                    money_nice = money
                    for i in range(len(spisok_variations)):
                        spisok_nice.append(spisok_variations[i])
            writing_to_a_file(name_file, spisok_variations, stroka_pls, water_expenditure, money)
            spisok_variations.clear()

            stroka_pls += 5
    return spisok_nice, money_nice


def writing_to_a_file(name_file,spisok_variations, i, water_expenditure, money):
    wb = load_workbook(name_file)
    sheet = wb.active
    first_device = 'A' + str(i)
    stolb_device = openpyxl.utils.cell.get_column_letter(len(spisok_variations)*4+1)
    second_device = stolb_device + str(i)
    obi = first_device + (':') + second_device
    megre_cell = sheet[first_device]
    megre_cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.merge_cells(obi)
    kolonk = 2
    sheet.cell(row=i + 2, column=kolonk - 1).value = 'Вещество'
    sheet.cell(row=i + 3, column=kolonk - 1).value = 'Взвешенные'
    end = 0
    eff = 1
    for j in range(len(spisok_variations)):
        first_device = openpyxl.utils.cell.get_column_letter(kolonk) + str(i + 1)
        stolb_device = openpyxl.utils.cell.get_column_letter(kolonk + 3)
        second_device = stolb_device + str(i + 1)
        obi = first_device + (':') + second_device
        megre_cell = sheet[first_device]
        megre_cell.alignment = Alignment(horizontal="center", vertical="center")
        sheet.merge_cells(obi)
        sheet.cell(row=i + 1, column=kolonk).value = spisok_variations[j].get('name')
        sheet.cell(row=i + 2, column=kolonk).value = 'C(нач)'
        sheet.cell(row=i + 2, column=kolonk + 1).value = 'n'
        sheet.cell(row=i + 2, column=kolonk + 2).value = 'C(кон)'
        sheet.cell(row=i + 2, column=kolonk + 3).value = 'C(кон)/ПДК'
        sheet.cell(row=i + 3, column=kolonk).value = spisok_variations[j].get('start')
        sheet.cell(row=i + 3, column=kolonk + 1).value = spisok_variations[j].get('effectiv')
        sheet.cell(row=i + 3, column=kolonk + 2).value = spisok_variations[j].get('end')
        sheet.cell(row=i + 3, column=kolonk + 3).value = spisok_variations[j].get('end/PDK')
        kolonk += 4
        end = j
        eff = eff * (1 - spisok_variations[j].get('effectiv')*0.01)
    sheet.cell(row=i + 1, column=kolonk).value = spisok_variations[end].get('общая информация по схеме')
    sheet.cell(row=i + 2, column=kolonk).value = 'C(нач. общ)'
    sheet.cell(row=i + 2, column=kolonk + 1).value = 'n(общ)'
    sheet.cell(row=i + 2, column=kolonk + 2).value = 'C(кон. общ)'
    sheet.cell(row=i + 2, column=kolonk + 3).value = 'C(кон)/ПДК'
    sheet.cell(row=i + 2, column=kolonk + 4).value = 'Стоимость'
    sheet.cell(row=i + 3, column=kolonk).value = spisok_variations[0].get('start')

    sheet.cell(row=i + 3, column=kolonk + 1).value = 1 - eff
    sheet.cell(row=i + 3, column=kolonk + 2).value = spisok_variations[end].get('end')
    sheet.cell(row=i + 3, column=kolonk + 3).value = spisok_variations[end].get('end') / water_expenditure
    sheet.cell(row=i + 3, column=kolonk + 4).value = money

    sheet.cell(row=i, column=1).value = 'Аппараты'

    wb.save(name_file)

def effectiveness(concent_changes, effectiveness):
    return concent_changes - (concent_changes * effectiveness / 100)



